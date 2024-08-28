# -*- coding: utf-8 -*-
"""
Created on Sun Feb 20 09:59:40 2022

@author: ilias
"""

import streamlit as st
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import pandas as pd
import os
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.styles.alignment import Alignment

st.set_page_config(layout="wide")
st.markdown("""# HEC-RAS output from .txt to pdf""")


# Helper function to create plots
def create_plots(data):
    data_plotting = data.copy()
    data_plotting.set_index('River Station', inplace=True)

    fig = make_subplots(rows=2, cols=1)
    trace_1 = go.Scatter(x=data_plotting.index, y=data_plotting['W.S. Elev'], name='water level (m)')
    trace_2 = go.Scatter(x=data_plotting.index, y=data_plotting['Min Ch El'], name='ground surface (m)')
    trace_3 = go.Scatter(x=data_plotting.index, y=data_plotting['Vel Chnl'], name='velocity (m/s)')

    fig.add_trace(trace_1, row=1, col=1)
    fig.add_trace(trace_2, row=1, col=1)
    fig.add_trace(trace_3, row=2, col=1)

    fig.update_layout(height=600, width=600, title_text="Hydraulic characteristics")
    return fig


# Helper function to save data to Excel
def save_to_excel(data, path):
    excel_name = os.path.join(path, 'Hydraulic.xlsx')
    data.to_excel(excel_name)
    return excel_name


# Helper function to adjust Excel formatting
def format_excel(excel_path):
    wb = load_workbook(excel_path)
    ws = wb['Sheet1']

    # Adjust column widths
    column_widths = []
    for row in ws.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))

    col_width_array = np.asarray(column_widths, dtype=np.float32) + 5
    for i, column_width in enumerate(col_width_array):
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width

    # Insert rows, freeze panes, and set alignment
    ws.insert_rows(1)
    ws.insert_rows(3)
    units_row = ['(m3/s)', '(m)', '(m)', '(m)', '(m)', '(m/m)', '(m/s)', '(m2)', '(m)', '-']
    for v, c in zip(units_row, range(5, 15)):
        ws.cell(column=c, row=3).value = v

    last_row = len(list(ws.rows))
    last_column = ws.max_column

    for j in range(1, last_column + 1):
        for i in range(2, last_row + 1):
            ws.cell(column=j, row=i).alignment = Alignment(horizontal='center')

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for j in range(1, last_column + 1):
        for i in range(2, last_row + 1):
            ws.cell(column=j, row=i).border = thin_border

    # Fix superscripts and set print settings
    ws['E3'] = '(m' + u'\u00B3' + '/s)'
    ws['L3'] = '(m' + u'\u00B2' + ')'

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False
    ws.print_title_rows = '1:3'

    wb.save(excel_path)


# Main application function
def main():
    col2, space2, col3 = st.columns((10, 1, 10))

    with col2:
        upload_data = st.file_uploader('Select the file', type=['txt'])
        if upload_data is not None:
            # Read and prepare data
            data = pd.read_fwf(upload_data, skiprows=10)
            col_names_English = ['Reach', 'River Station', 'Profile', 'Number', 'Q Total', 'Min Ch El', 'W.S. Elev',
                                 'Crit W.S.', 'E.G. Elev', 'E.G. Slope', 'Vel Chnl', 'Flow Area', 'Top Width',
                                 'Froude # Chl']
            data.columns = col_names_English

            data['Number'] = data['Number'].astype('str')
            data['Profile'] = data['Profile'] + data['Number']
            data.drop('Number', axis=1, inplace=True)

            # Create and display plots
            fig = create_plots(data)
            st.plotly_chart(fig, use_container_width=True)

            # Create directory
            cwd = os.getcwd()
            directory_name = 'HEC_RAS_excel_pdf'
            path = os.path.join(cwd, directory_name)
            if not os.path.exists(path):
                os.mkdir(path)

            # Save data to Excel
            excel_path = save_to_excel(data, path)
            format_excel(excel_path)

            with col3:
                selected_cat = st.selectbox(label='Select parameter', options=['Only Excel file'])

                with open(excel_path, "rb") as file:
                    st.download_button(
                        label="Click to download EXCEL",
                        data=file,
                        file_name="downloaded_EXCEL.xlsx",
                        mime="application/vnd.ms-excel"
                    )


# Run the application
if __name__ == "__main__":
    main()
